"""Builds dashboard/PhonePe_Digital_Payments_Dashboard.xlsx from outputs/*.csv."""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "outputs"
DASH = ROOT / "dashboard"
DASH.mkdir(exist_ok=True)


def write_df(ws, df, start_row=1):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 30)


wb = Workbook()
wb.properties.creator = "Aryan Ahlawat"
wb.properties.lastModifiedBy = "Aryan Ahlawat"
wb.properties.title = "Digital Payments Transaction Analytics"

# --- Sheet 1: daily volume pivot (day x type) ---
daily = pd.read_csv(OUT / "daily_volume_by_type.csv")
pivot = daily.pivot_table(index="day", columns="type", values="txn_value", fill_value=0).reset_index()
ws1 = wb.active
ws1.title = "Daily Volume Pivot"
write_df(ws1, pivot)

# --- Sheet 2: forecast + line chart ---
fc = pd.read_csv(OUT / "daily_volume_with_forecast.csv")
fc_wide = fc.pivot_table(index="day", columns="type", values="value").reset_index()
ws2 = wb.create_sheet("Forecast")
write_df(ws2, fc_wide)

chart1 = LineChart()
chart1.title = "Daily Transaction Value: Actual vs Forecast"
chart1.y_axis.title = "txn value"
chart1.x_axis.title = "day"
data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=ws2.max_row)
cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws2.add_chart(chart1, "F2")

# --- Sheet 3: fraud + segments + 2 charts ---
fraud = pd.read_csv(OUT / "fraud_rate_by_type.csv")
seg = pd.read_csv(OUT / "segment_summary.csv")
ws3 = wb.create_sheet("Fraud & Segments")
ws3.append(["Fraud rate by type"])
write_df(ws3, fraud, start_row=2)
fraud_end_row = ws3.max_row

ws3.append([])
seg_start_row = ws3.max_row + 1
ws3.append(["Segment summary"])
write_df(ws3, seg)

chart2 = BarChart()
chart2.title = "Fraud Rate by Type (%)"
chart2.y_axis.title = "fraud_rate_pct"
data = Reference(ws3, min_col=4, min_row=2, max_row=fraud_end_row)
cats = Reference(ws3, min_col=1, min_row=3, max_row=fraud_end_row)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws3.add_chart(chart2, "H2")

chart3 = PieChart()
chart3.title = "Value Share by Segment"
seg_header_row = seg_start_row + 1
seg_end_row = seg_header_row + len(seg)
data = Reference(ws3, min_col=3, min_row=seg_header_row, max_row=seg_end_row)
cats = Reference(ws3, min_col=1, min_row=seg_header_row + 1, max_row=seg_end_row)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
ws3.add_chart(chart3, "H18")

# --- Sheet 4: other query results ---
ws4 = wb.create_sheet("Other Query Results")
top_senders = pd.read_csv(OUT / "top_senders.csv")
balance = pd.read_csv(OUT / "balance_check.csv")
fraud_chains = pd.read_csv(OUT / "fraud_chains.csv")

ws4.append(["Top senders (by total value sent)"])
write_df(ws4, top_senders, start_row=2)
ws4.append([])
ws4.append(["Balance reconciliation check (oldbalanceOrg - amount != newbalanceOrig)"])
write_df(ws4, balance)
ws4.append([])
ws4.append(["Fraud chain search (TRANSFER->CASH_OUT via shared account, <=3 steps, <5% amount drift)"])
if len(fraud_chains) == 0:
    ws4.append(["No matches in full dataset. See README for the negative-result writeup."])
else:
    write_df(ws4, fraud_chains)

wb.save(DASH / "PhonePe_Digital_Payments_Dashboard.xlsx")
print(f"wrote {DASH / 'PhonePe_Digital_Payments_Dashboard.xlsx'}")
